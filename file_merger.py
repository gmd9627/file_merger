import os
import pandas as pd
from openpyxl import load_workbook, Workbook

# ✅ Define file paths
file1 = "/Users/garrettdeese/Tech_Drivers/00-Six-Pack-Analysis-Template.xlsx"
file2 = "/Users/garrettdeese/Tech_Drivers/Adobe Inc NasdaqGS ADBE Financials.xlsx"
output_file = "/Users/garrettdeese/Tech_Drivers/sixpackmerge.xlsx"

# ✅ Check if files exist before proceeding
for file in [file1, file2]:
    if not os.path.exists(file):
        raise FileNotFoundError(f"❌ File not found: {file}")

# ✅ Load both Excel workbooks
wb1 = load_workbook(file1, data_only=True)  # Read values instead of formulas
wb2 = load_workbook(file2, data_only=True)

# ✅ Create a new workbook for the merged file
merged_wb = Workbook()
merged_wb.remove(merged_wb.active)  # Remove default empty sheet

# ✅ Function to copy sheets while safely handling styles
def copy_sheets(source_wb, target_wb):
    for sheet_name in source_wb.sheetnames:
        source_ws = source_wb[sheet_name]

        # ✅ Skip empty sheets
        if source_ws.max_row == 1 and source_ws.max_column == 1 and source_ws["A1"].value is None:
            print(f"⚠️ Skipping empty sheet: {sheet_name}")
            continue

        print(f"✅ Copying sheet: {sheet_name}")
        target_ws = target_wb.create_sheet(title=sheet_name)

        for row in source_ws.iter_rows():
            for cell in row:
                target_ws[cell.coordinate] = cell.value  # Copy values

                # ✅ Skip problematic styles completely
                try:
                    if cell.has_style and hasattr(cell, "_style"):
                        target_ws[cell.coordinate]._style = cell._style
                except (IndexError, AttributeError, KeyError):
                    print(f"⚠️ Skipping corrupted style in {sheet_name} at {cell.coordinate}")

        # ✅ Preserve column widths
        for col in source_ws.column_dimensions:
            if col in source_ws.column_dimensions and col in target_ws.column_dimensions:
                target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width

# ✅ Copy sheets from both workbooks
copy_sheets(wb1, merged_wb)
copy_sheets(wb2, merged_wb)

# ✅ Ensure the merged file contains data before saving
if len(merged_wb.sheetnames) == 0:
    raise ValueError("❌ No sheets were copied! Check that the original files have data.")

# ✅ Save the final merged workbook
try:
    merged_wb.save(output_file)
    print(f"✅ Successfully merged with formatting into {output_file}")
except Exception as e:
    print(f"❌ Error saving the file: {e}")
